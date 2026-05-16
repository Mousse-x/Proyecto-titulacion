"""
Script para analizar los 25 archivos XLSX de plantillas oficiales LOTAIP
y pre-configurar expected_columns y keywords en IndicatorTemplate.

Ejecutar con PYTHONIOENCODING=utf-8
"""
import os
import re
import sys
import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

import openpyxl
from pathlib import Path
from django.conf import settings
from core.models import IndicatorTemplate


def normalize_col(text):
    if not text:
        return ""
    text = str(text).strip()
    text = text.replace('\n', ' ').replace('\r', ' ')
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def extract_columns_from_xlsx(file_path):
    columns = []
    keywords = []
    structure = {}

    try:
        wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
        structure["sheets"] = wb.sheetnames

        # 1. Buscar columnas en la hoja "Conjunto de datos" (nombre estándar LOTAIP)
        data_sheet = None
        for name in wb.sheetnames:
            if 'conjunto' in name.lower() or 'datos' in name.lower():
                data_sheet = wb[name]
                break

        if data_sheet is None:
            # Usar la primera hoja si no hay "Conjunto de datos"
            data_sheet = wb[wb.sheetnames[0]]

        # La primera fila de "Conjunto de datos" tiene los encabezados reales
        rows = list(data_sheet.iter_rows(max_row=5, values_only=True))
        if rows:
            first_row = rows[0]
            columns = [normalize_col(str(c)) for c in first_row
                       if c is not None and normalize_col(str(c))]

        # Contar filas totales
        total_rows = data_sheet.max_row or 0
        structure["min_rows"] = max(2, total_rows // 2)
        structure["total_rows_template"] = total_rows

        # 2. Buscar keywords en hoja "Diccionario" o "Metadatos"
        for name in wb.sheetnames:
            sheet = wb[name]
            for row in sheet.iter_rows(max_row=10, values_only=True):
                for cell in row:
                    if cell and isinstance(cell, str) and len(cell.strip()) > 3:
                        val = cell.strip()
                        if any(kw in val.lower() for kw in [
                            'lotaip', 'transparencia', 'art.', 'literal',
                            'periodicidad', 'mensual', 'anual'
                        ]):
                            keywords.append(val[:100])

        wb.close()

    except Exception as e:
        print(f"  Error procesando {file_path}: {e}")

    # Limpiar columnas
    clean_columns = [c for c in columns if c and len(c) > 1]

    return clean_columns, list(set(keywords)), structure


def main():
    templates = IndicatorTemplate.objects.select_related('indicator').all()

    if templates.count() == 0:
        print("No hay plantillas en la base de datos.")
        return

    updated = 0

    for template in templates:
        file_path = Path(settings.MEDIA_ROOT) / str(template.file_path)

        if not file_path.exists():
            print(f"[WARN] {template.indicator.code}: Archivo no encontrado -> {file_path}")
            continue

        print(f"[+] Procesando: {template.indicator.code} - {template.file_name}")

        columns, keywords, structure = extract_columns_from_xlsx(file_path)

        if columns:
            template.expected_columns = columns
            template.keywords = keywords
            template.expected_structure = structure
            template.save(update_fields=['expected_columns', 'keywords', 'expected_structure'])
            updated += 1
            print(f"    OK: {len(columns)} columnas, {len(keywords)} keywords")
            for c in columns:
                print(f"      - {c}")
        else:
            print(f"    WARN: No se encontraron columnas")

    print(f"\n{'='*60}")
    print(f"OK: {updated}/{templates.count()} plantillas configuradas.")
    print(f"{'='*60}")


main()
